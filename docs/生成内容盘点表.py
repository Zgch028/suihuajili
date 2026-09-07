#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成公众号内容盘点表模板（xlsx）。
用法：python docs/生成内容盘点表.py
输出：docs/公众号内容盘点表.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "公众号内容盘点"

headers = [
    "序号",
    "公众号合集名",
    "对应网站 series slug",
    "文章标题",
    "发布日期",
    "原文链接",
    "是否需要搬运",
    "备注",
]

ws.append(headers)

# 表头样式
header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 示例数据
examples = [
    [1, "大宋纪丽", "dasongjili", "点茶不是泡茶：从一盏建盏看宋人的慢", "2026-06-18", "https://mp.weixin.qq.com/...", "是", "已搬运"],
    [2, "十二生肖职场人格解码", "shiershengxiao", "像狗一样守护", "2026-08-17", "https://mp.weixin.qq.com/...", "是", "已搬运"],
    [3, "二十四节气", "ershisijieqi", "", "", "", "", ""],
    [4, "一宋一词", "yisongyici", "", "", "", "", ""],
    [5, "纪丽生活馆", "jilishenghuoguan", "", "", "", "", ""],
    [6, "元一·钻石监理成长圈", "yuanyizuanshi", "", "", "", "", ""],
    [7, "视频号", "shipinhao", "", "", "", "", ""],
]

for row in examples:
    ws.append(row)

# 列宽
column_widths = [6, 22, 24, 40, 12, 32, 12, 20]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 边框
thin = Side(style="thin", color="d1d5db")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 冻结首行
ws.freeze_panes = "A2"

output_path = "docs/公众号内容盘点表.xlsx"
wb.save(output_path)
print(f"已生成：{output_path}")
